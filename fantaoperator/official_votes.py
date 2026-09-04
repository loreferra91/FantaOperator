"""Original, fail-closed parser for vote exports. No third-party scraper code."""
from __future__ import annotations

import io
import json
import math
import re
from dataclasses import dataclass
from typing import Any, Iterable, Mapping
from zipfile import ZipFile, BadZipFile

from .sources import MAX_PAYLOAD, normalize_key, normalized_fields, load_json, parse_payload

EDITIONS = ("Redazione Fantacalcio", "Voto Statistico", "Voto Italia")
STATUSES = ("LIVE", "PROVVISORIO", "DEFINITIVO")
COUNTERS = (
    "goals", "assists", "yellow_cards", "red_cards", "own_goals",
    "goals_conceded", "penalties_saved", "penalties_scored", "penalties_missed",
)
SCORING_FIELDS = (*COUNTERS, "clean_sheet", "custom_bonus", "custom_malus")
NO_VOTE = {"sv", "s.v.", "s.v", "s/v", "-", "--"}
AMBIGUOUS_FIELDS = {"rf", "rs", "rc", "rig", "r+", "r-", "assf", "asss"}
ACCEPTED_FIELDS = {
    "player", "vote", "role", "team", "status", "provider_player_id", "provider",
    "edition", "season", "matchday", "schema_version", "fantavote", "fv",
    *SCORING_FIELDS, *AMBIGUOUS_FIELDS,
}


def season_name(value: Any) -> str:
    match = re.fullmatch(r"(20\d{2})[-/](\d{2}|20\d{2})", str(value).strip())
    if not match or int(match[2][-2:]) != (int(match[1]) + 1) % 100:
        raise ValueError("Stagione non valida: usa ad esempio 2026-27")
    return f"{match[1]}-{match[2][-2:]}"


def number(value: Any, field: str) -> float:
    try:
        if isinstance(value, bool):
            raise ValueError
        result = float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError):
        raise ValueError(f"Valore numerico non valido: {field}") from None
    if not math.isfinite(result):
        raise ValueError(f"Valore non finito: {field}")
    return result


def normalize_rows(rows: Iterable[Mapping[str, Any]], default_status: str = "PROVVISORIO") -> list[dict]:
    if default_status not in STATUSES:
        raise ValueError("Stato dati non valido")
    result, seen = [], set()
    for raw in rows:
        if not isinstance(raw, Mapping):
            raise ValueError("Ogni record deve essere un oggetto")
        item = normalized_fields(raw)
        unknown = set(item) - ACCEPTED_FIELDS
        if unknown:
            raise ValueError(f"Colonne non riconosciute: {', '.join(sorted(unknown))}. Usa il modello CSV/JSON.")
        if not isinstance(item.get("player"), str):
            raise ValueError("Il nome giocatore deve essere testo")
        name = str(item.get("player") or "").strip()
        if not name or len(name) > 150:
            raise ValueError("Nome giocatore mancante o troppo lungo")
        if name.casefold() in seen:
            raise ValueError(f"Giocatore duplicato/omonimo da risolvere: {name}")
        seen.add(name.casefold())
        if "vote" not in item or item["vote"] == "":
            raise ValueError(f"Colonna voto mancante: {name}")
        raw_vote = item["vote"]
        vote = None if raw_vote is None or str(raw_vote).strip().lower() in NO_VOTE else number(raw_vote, "voto")
        if vote is not None and not 0 <= vote <= 10:
            raise ValueError(f"Voto fuori intervallo 0–10: {name}")
        status = str(item.get("status") or default_status).strip().upper()
        if status not in STATUSES:
            raise ValueError(f"Stato dati non valido: {name}")
        role = str(item.get("role") or "").strip().upper()
        role = {"P": "POR", "D": "DIF", "C": "CEN", "A": "ATT"}.get(role, role)
        row = {"player": name, "vote": vote, "status": status, "role": role,
               "team": str(item.get("team") or "").strip()}
        for key in COUNTERS:
            value = number(0 if item.get(key) in (None, "") else item[key], key)
            if not value.is_integer() or not 0 <= value <= 100:
                raise ValueError(f"Conteggio non valido: {key} ({name})")
            row[key] = int(value)
        # Goals includes penalty goals; the scoring engine must never count them twice.
        if row["penalties_scored"] > row["goals"]:
            raise ValueError("I rigori segnati devono essere inclusi nei gol totali")
        clean = str(item.get("clean_sheet") or "0").lower()
        if clean not in {"0", "1", "true", "false", "si", "sì", "no", "yes"}:
            raise ValueError("Clean sheet non valido")
        row["clean_sheet"] = None if "clean_sheet" in item and item["clean_sheet"] is None else clean in {"1", "true", "si", "sì", "yes"}
        for key in ("custom_bonus", "custom_malus"):
            row[key] = number(item.get(key) or 0, key)
        for key in AMBIGUOUS_FIELDS:
            if item.get(key) not in (None, "", 0, "0"):
                raise ValueError(f"Colonna ambigua {key}: usa intestazioni esplicite del modello JSON/CSV")
        result.append(row)
        if len(result) > 2000:
            raise ValueError("Troppi record: limite 2000 giocatori")
    if not result:
        raise ValueError("Nessun voto disponibile nella sorgente")
    return result


def _xlsx_rows(payload: bytes, edition: str) -> tuple[list[dict], str]:
    from openpyxl import load_workbook

    try:
        with ZipFile(io.BytesIO(payload)) as archive:
            if len(archive.infolist()) > 500 or sum(x.file_size for x in archive.infolist()) > 30 * 1024 * 1024:
                raise ValueError("XLSX decompresso troppo grande")
    except BadZipFile:
        raise ValueError("File XLSX non valido") from None
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False, keep_links=False)
    try:
        # Multiple editorial sheets are never silently merged or selected by position.
        matching = [s for s in workbook.worksheets if s.title.casefold() == edition.casefold()]
        if len(workbook.worksheets) > 1 and not matching:
            raise ValueError("XLSX con più fogli: il nome del foglio deve coincidere con la redazione configurata")
        sheet = matching[0] if matching else workbook.worksheets[0]
        if sheet.max_column and sheet.max_column > 60:
            raise ValueError("XLSX con troppe colonne: limite 60, nessuna colonna viene troncata")
        if sheet.title in EDITIONS and sheet.title != edition:
            raise ValueError("Redazione XLSX diversa da quella configurata")
        rows, headers, team, preamble = [], None, "", []
        for index, values in enumerate(sheet.iter_rows(max_col=60, values_only=True)):
            if index > 6000:
                raise ValueError("XLSX troppo lungo")
            cells = list(values)
            if any(isinstance(v, str) and v.startswith("=") for v in cells):
                raise ValueError("XLSX con formule: esporta valori statici, senza formule")
            nonempty = [str(v).strip() for v in cells if v is not None and str(v).strip()]
            if not nonempty:
                continue
            keys = [normalize_key(v) if v is not None else "" for v in cells]
            if "player" in keys and "vote" in keys:
                used = [k for k in keys if k]
                if len(used) != len(set(used)):
                    raise ValueError("Intestazioni XLSX duplicate: seleziona una sola redazione")
                headers = keys
                continue
            if len(nonempty) == 1:
                team = nonempty[0]
                if index < 5:
                    preamble.append(team)
                continue
            if headers is None:
                if index < 10:
                    preamble.extend(nonempty)
                continue
            row = {key: value for key, value in zip(headers, cells) if key}
            if not row.get("player"):
                raise ValueError("Riga XLSX non riconosciuta dopo l'intestazione")
            # Fantacalcio's coach code is not a Classic attacking player.
            if str(row.get("role", "")).upper() in {"ALL", "ALLENATORE"} or ("provider_player_id" in row and row.get("role") == "ATT"):
                continue
            row.setdefault("team", team)
            rows.append(row)
        return rows, " ".join(preamble)
    finally:
        workbook.close()


@dataclass(frozen=True)
class VoteBatch:
    records: list[dict]
    warnings: tuple[str, ...]


def parse_votes(payload: bytes, filename: str, content_type: str, *, provider: str,
                edition: str, season: str, matchday: int, remote: bool = False,
                default_status: str = "PROVVISORIO") -> VoteBatch:
    if len(payload) > MAX_PAYLOAD:
        raise ValueError("File troppo grande: limite 5 MB")
    season = season_name(season)
    if not 1 <= matchday <= 38:
        raise ValueError("Giornata non valida")
    metadata, preamble = {}, ""
    is_xlsx = filename.lower().split("?")[0].endswith(".xlsx") or payload.startswith(b"PK\x03\x04")
    if is_xlsx:
        rows, preamble = _xlsx_rows(payload, edition)
    else:
        if payload.decode("utf-8-sig").lstrip().startswith("{"):
            data = load_json(payload)
            metadata = {k: data[k] for k in ("provider", "edition", "season", "matchday", "schema_version") if k in data}
            if metadata.get("schema_version", 1) != 1:
                raise ValueError("Versione schema JSON non supportata")
        rows = parse_payload(payload, filename, content_type)
    # Canonical metadata can also be repeated on every CSV record.
    expected = {"provider": provider, "edition": edition, "season": season, "matchday": matchday}
    for key in expected:
        values = {str(row[key]) for row in rows if row.get(key) not in (None, "")}
        if len(values) > 1:
            raise ValueError(f"Metadati discordanti: {key}")
        if values:
            row_value = values.pop()
            if key in metadata and str(metadata[key]) != row_value:
                raise ValueError(f"Metadati discordanti: {key}")
            metadata[key] = row_value
    # Read period claims from official-style filenames/titles, never infer from today's date.
    identity = filename + " " + preamble
    season_match = re.search(r"20\d{2}[-/]\d{2}(?:\d{2})?", identity)
    day_match = re.search(r"giornata[_\s-]*(\d{1,2})|\b(\d{1,2})[ª°]?\s*giornata", identity, re.I)
    claims = dict(metadata)
    for key, value in (("season", season_match[0] if season_match else None),
                       ("matchday", (day_match[1] or day_match[2]) if day_match else None)):
        if value is not None:
            if key in claims and str(claims[key]).replace("/", "-") != str(value).replace("/", "-"):
                raise ValueError(f"Periodo discordante nel file: {key}")
            claims[key] = value
    for key, expected_value in expected.items():
        if key not in claims:
            continue
        actual = season_name(claims[key]) if key == "season" else str(claims[key]).strip()
        if actual != str(expected_value):
            raise ValueError(f"Fonte/periodo non corrispondente alla lega: {key}")
    # Automated feeds must declare all four fields. A URL alone proves no editorial provenance.
    if remote and any(key not in claims for key in expected):
        raise ValueError("Feed automatico privo di provider, edition, season o matchday. Usa il JSON normalizzato documentato.")
    warnings = () if remote else ("Import locale: provenienza e stato dichiarati dall'utente, non verificati sul Web.",)
    return VoteBatch(normalize_rows(rows, default_status), warnings)
